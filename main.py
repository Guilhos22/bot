
"""
This script was created with the idea of automating the process of uploading CT-es (Electronic Transport Documents) online in the department where I work.
The code will remain open, with some parts replaced, such as the website used for submission, for example.

If you have suggestions for improvements, feel free to contact me below :)

Creator: Guilherme Henrique Alves De Oliveira  
LinkedIn: https://www.linkedin.com/in/guilherme-henrique-205083316?utm_source=share&utm_campaign=share_via&utm_content=profile&utm_medium=ios_app  
GitHub: https://github.com/Guilhos22  
Email/Contact: guilherme@guilhoslabs.com.br
best regards,
Guilherme Henrique! 

"""
import pyperclip
import sys
import openpyxl
from interface import iniciar_interface, mostrar_erro_popup, mostrar_popup


def main():
    cte_list = []
    while True: 
        branch, data_in = iniciar_interface()

        if not branch or not data_in:
            sys.exit(0)

        def excel_list(sheet):
            cte_list.clear()
            for cte in sheet.iter_rows(min_row=2):
                cte_value = str(cte[0].value).strip()
                if cte_value:
                    cte_list.append(cte_value)
            if not cte_list:
                mostrar_erro_popup("Error", f'A filial: "{branch}" selecionada está vazia')
                sys.exit(1)
            #for v in cte_list:      # v == value
                #pyperclip.copy(v)
                    #print(v) debug
        try:
            excel_file_name = "data_base.xlsx".lower().strip()
            excel = openpyxl.load_workbook(excel_file_name)
            sheet_xml = excel[branch]
            excel_list(sheet_xml)
        except FileNotFoundError:
            mostrar_erro_popup("Excel Error", f"Não foi possivel encontrar nenhuma planilha\n\n '{excel_file_name}' ")
            from interface import criar_excel
            if criar_excel():
                try:
                    wb = openpyxl.Workbook()
                    work_sheet_1 = wb.active
                    work_sheet_1.title = "aromas"
                    work_sheet_2 = wb.create_sheet("matriz")
                    work_sheet_3 = wb.create_sheet("produção")

                    work_sheet_1['A1'] = "Comece a partir de 'A2'"
                    work_sheet_2['A1'] = "Comece a partir de 'A2'"
                    work_sheet_3['A1'] = "Comece a partir de 'A2'"

                    wb.save(f"{excel_file_name}")
                    mostrar_popup("Sucesso!", "Excel criado com sucesso!\n\n Lembre-se de preencher cada filial")
                    sys.exit(0)
                except Exception as error:
                    mostrar_erro_popup("Error", f"Um erro inesperado aconteceu {error}")
                    continue
            else:
                sys.exit(1)
        except Exception as generic_error:
            mostrar_erro_popup("Excel Error", f"Grave erro. Tente novamente mais tarde ou entre com contato com adm: guilherme@guilhoslabs.com.br\n\n {generic_error} ")
            sys.exit(1)

        if branch == "matriz":
            from branch import fragrance
            fragrance(data_in,cte_list)

        elif branch == "aromas":
            from branch import aromas
            aromas(data_in,cte_list)

        elif branch == "produção":
            from branch import producao
            producao(data_in,cte_list)

        from interface import continuar
        if not continuar(): 
            sys.exit(0) # closing program without no error 

if __name__ == "__main__":
    main()